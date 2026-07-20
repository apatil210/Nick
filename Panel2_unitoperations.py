from io import BytesIO
import re

import pandas as pd
import plotly.express as px
import plotly.io as pio
import requests
import streamlit as st
from openpyxl import load_workbook

# ----------------------------
# App configuration
# ----------------------------
st.set_page_config(
    page_title="US Manufacturing Energy Classification",
    layout="wide"
)

pio.templates.default = "plotly"

DATA_URL = "https://raw.githubusercontent.com/apatil210/Nick/main/WebsiteEngine3.xlsx"
SHEET_NAME = "Process-level data"

TEXT_COLOR = "#14212B"
PAPER_BG = "rgba(0,0,0,0)"
PLOT_BG = "rgba(0,0,0,0)"
BAR_COLOR = "#0B6E74"

ENERGY_COLOR_MAP = {
    "Electricity": "#54A24B",
    "Fuels": "#F58518",
    "Steam": "#4C78A8",
}

TEMP_COLOR_MAP_SI = {
    "<20 °C": "#72B7B2",
    "20-100 °C": "#54A24B",
    "100-200 °C": "#EACA3B",
    "200-400 °C": "#ECA82C",
    "400-600 °C": "#F58518",
    ">=600 °C": "#E45756",
}

TEMP_COLOR_MAP_IMPERIAL = {
    "<68 °F": "#72B7B2",
    "68-212 °F": "#54A24B",
    "212-392 °F": "#EACA3B",
    "392-752 °F": "#ECA82C",
    "752-1112 °F": "#F58518",
    ">=1112 °F": "#E45756",
}

TEMP_LABEL_MAP_SI_TO_IMPERIAL = {
    "<20 °C": "<68 °F",
    "20-100 °C": "68-212 °F",
    "100-200 °C": "212-392 °F",
    "200-400 °C": "392-752 °F",
    "400-600 °C": "752-1112 °F",
    ">=600 °C": ">=1112 °F",
}

UNIT_CONFIG = {
    "SI": {
        "annual_energy_unit": "PJ/yr",
        "sec_unit": "GJ/t",
        "temp_unit": "°C",
        "pressure_unit": "bar",
        "temp_color_map": TEMP_COLOR_MAP_SI,
    },
    "Imperial": {
        "annual_energy_unit": "TBtu/yr",
        "sec_unit": "MMBtu/short ton",
        "temp_unit": "°F",
        "pressure_unit": "psi",
        "temp_color_map": TEMP_COLOR_MAP_IMPERIAL,
    }
}

# ----------------------------
# Exact workbook column names
# ----------------------------
COL_L2 = "Unit operation (Level 2 classification)"
COL_L3 = "Industrial process"
COL_PERCENT_ENERGY = "Percent Annual energy demand in 2022"
COL_PERCENT_ELECTRICITY = "Percent Annual electricity demand in 2022"
COL_PERCENT_FUELS = "Perent Annual fuels demand in 2022"
COL_PERCENT_STEAM = "Percent Annual fuels or electricity for steam or steam from CHP demand in 2022"

COL_ANNUAL_PRODUCTION = "Annual production in 2022\n(based on FU)"

COL_SEC_ELECTRICITY = "SEC electricity"
COL_SEC_FUELS = "SEC fuels"
COL_SEC_STEAM = "SEC fuels or electricity for steam or steam from CHP"

COL_EFFICIENCY = "Efficiency"
COL_PROCESS_TEMP = "Process temperature"
COL_PROCESS_TEMP_WEB = "Process Temperature for Webpage"
COL_INLET_TEMP = "Inlet temperature"
COL_OUTLET_TEMP = "Outlet temperature"
COL_PROCESS_PRESSURE = "Process pressure"
COL_INLET_PRESSURE = "Inlet pressure"
COL_OUTLET_PRESSURE = "Outlet pressure"
COL_RESIDENCE_TIME = "Residence time"
COL_NAICS = "NAICS Level 2 Code Number"

# ----------------------------
# Exact Excel columns (1-based)
# ----------------------------
EXCEL_COL_E = 5
EXCEL_COL_K = 11
EXCEL_COL_AU = 47
EXCEL_COL_AW = 49
EXCEL_COL_AX = 50
EXCEL_COL_AY = 51

# ----------------------------
# Unit conversions
# ----------------------------
def c_to_f(x):
    return x * 9 / 5 + 32


def bar_to_psi(x):
    return x * 14.5038


def pj_to_tbtu(x):
    return x * 0.947817


def gj_per_t_to_mmbtu_per_short_ton(x):
    return x * 0.947817 / 1.10231

# ----------------------------
# Header utilities
# ----------------------------
def normalize_header(value) -> str:
    text = str(value if value is not None else "")
    text = text.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def canonical_header(value) -> str:
    text = normalize_header(value).lower()
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def find_matching_column(df: pd.DataFrame, target: str) -> str:
    target_canonical = canonical_header(target)

    for col in df.columns:
        if canonical_header(col) == target_canonical:
            return col

    available = "\n".join([f"- {repr(col)}" for col in df.columns])
    raise KeyError(
        f"Could not find required column: {target}\nAvailable columns:\n{available}"
    )

# ----------------------------
# Cleaning utilities
# ----------------------------
def clean_category(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def clean_naics(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def clean_text(series: pd.Series) -> pd.Series:
    return (
        series.astype(str)
        .str.replace("\xa0", " ", regex=False)
        .str.strip()
        .replace({"": "Unknown", "nan": "Unknown", "None": "Unknown"})
    )


def to_numeric_safe(series: pd.Series) -> pd.Series:
    cleaned = (
        series.astype(str)
        .str.replace("\xa0", "", regex=False)
        .str.replace(",", "", regex=False)
        .str.replace("%", "", regex=False)
        .str.strip()
    )
    return pd.to_numeric(cleaned, errors="coerce")

# ----------------------------
# Data loading
# ----------------------------
@st.cache_data(show_spinner=False)
def load_excel_data(url: str) -> pd.DataFrame:
    response = requests.get(url, timeout=30)
    response.raise_for_status()

    content_type = response.headers.get("Content-Type", "")
    if "text/html" in content_type.lower():
        raise ValueError("The URL returned HTML instead of an Excel file.")

    wb = load_workbook(filename=BytesIO(response.content), data_only=True)
    ws = wb[SHEET_NAME]

    header_row_excel = 2
    data_start_excel = 4

    max_col = ws.max_column
    max_row = ws.max_row

    headers = [
        ws.cell(row=header_row_excel, column=c).value
        for c in range(1, max_col + 1)
    ]

    rows = []
    for r in range(data_start_excel, max_row + 1):
        row_values = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None for v in row_values):
            continue
        rows.append(row_values)

    df = pd.DataFrame(
        rows,
        columns=[str(h) if h is not None else f"Unnamed_{i+1}" for i, h in enumerate(headers)]
    )

    df[COL_PROCESS_TEMP_WEB] = pd.Series([row[EXCEL_COL_K - 1] for row in rows])
    df["_Description_Excel_E"] = pd.Series([row[EXCEL_COL_E - 1] for row in rows])
    df["_Annual_Energy_AU"] = pd.Series([row[EXCEL_COL_AU - 1] for row in rows])
    df["_Annual_Electricity_AW"] = pd.Series([row[EXCEL_COL_AW - 1] for row in rows])
    df["_Annual_Fuels_AX"] = pd.Series([row[EXCEL_COL_AX - 1] for row in rows])
    df["_Annual_Steam_AY"] = pd.Series([row[EXCEL_COL_AY - 1] for row in rows])

    return df

# ----------------------------
# Data preparation
# ----------------------------
def prepare_treemap_data(df: pd.DataFrame, metric_col_name: str, top_n: int = 10) -> pd.DataFrame:
    l2_col = find_matching_column(df, COL_L2)
    metric_col = find_matching_column(df, metric_col_name)

    working_df = df[[l2_col, metric_col]].copy()
    working_df[l2_col] = clean_category(working_df[l2_col])
    working_df[metric_col] = to_numeric_safe(working_df[metric_col]).fillna(0)

    grouped_df = (
        working_df.groupby(l2_col, as_index=False)[metric_col]
        .sum()
        .sort_values(metric_col, ascending=False)
        .reset_index(drop=True)
    )

    grouped_df = grouped_df[grouped_df[metric_col] != 0].copy()

    max_abs_val = grouped_df[metric_col].abs().max()
    if pd.notna(max_abs_val) and max_abs_val <= 1:
        grouped_df["Display Percent"] = grouped_df[metric_col] * 100
    else:
        grouped_df["Display Percent"] = grouped_df[metric_col]

    grouped_df = grouped_df.rename(columns={
        l2_col: "Unit operation (Level 2 classification)",
        metric_col: "Raw Metric"
    })

    grouped_df = grouped_df.nlargest(top_n, "Display Percent").reset_index(drop=True)
    grouped_df["Rank"] = range(1, len(grouped_df) + 1)

    return grouped_df


def build_fact_sheet(df: pd.DataFrame, selected_l2: str):
    l2_col = find_matching_column(df, COL_L2)
    l3_col = find_matching_column(df, COL_L3)
    annual_prod_col = find_matching_column(df, COL_ANNUAL_PRODUCTION)

    sec_elec_col = find_matching_column(df, COL_SEC_ELECTRICITY)
    sec_fuels_col = find_matching_column(df, COL_SEC_FUELS)
    sec_steam_col = find_matching_column(df, COL_SEC_STEAM)

    efficiency_col = find_matching_column(df, COL_EFFICIENCY)
    process_temp_col = find_matching_column(df, COL_PROCESS_TEMP)
    inlet_temp_col = find_matching_column(df, COL_INLET_TEMP)
    outlet_temp_col = find_matching_column(df, COL_OUTLET_TEMP)
    process_pressure_col = find_matching_column(df, COL_PROCESS_PRESSURE)
    inlet_pressure_col = find_matching_column(df, COL_INLET_PRESSURE)
    outlet_pressure_col = find_matching_column(df, COL_OUTLET_PRESSURE)
    residence_time_col = find_matching_column(df, COL_RESIDENCE_TIME)
    naics_col = find_matching_column(df, COL_NAICS)

    annual_energy_col = "_Annual_Energy_AU"
    annual_elec_col = "_Annual_Electricity_AW"
    annual_fuels_col = "_Annual_Fuels_AX"
    annual_steam_col = "_Annual_Steam_AY"
    description_col = "_Description_Excel_E"
    process_temp_web_col = COL_PROCESS_TEMP_WEB

    fact_df = df.copy()
    fact_df[l2_col] = clean_category(fact_df[l2_col])
    fact_df[l3_col] = clean_category(fact_df[l3_col])
    fact_df[naics_col] = clean_naics(fact_df[naics_col])

    selected_l2_clean = str(selected_l2).replace("\xa0", " ").strip()
    selected_df = fact_df[fact_df[l2_col] == selected_l2_clean].copy()

    if selected_df.empty:
        return None

    numeric_cols = [
        annual_prod_col,
        annual_energy_col,
        annual_elec_col,
        annual_fuels_col,
        annual_steam_col,
        sec_elec_col,
        sec_fuels_col,
        sec_steam_col,
        process_temp_col,
        process_temp_web_col,
        process_pressure_col,
        inlet_pressure_col,
        outlet_pressure_col,
        inlet_temp_col,
        outlet_temp_col,
        residence_time_col,
    ]

    for col in numeric_cols:
        selected_df[col] = to_numeric_safe(selected_df[col])

    production_values = (
        selected_df[annual_prod_col]
        .dropna()
        .loc[lambda s: s != 0]
        .unique()
    )
    annual_production = production_values[0] if len(production_values) > 0 else 0

    annual_energy = selected_df[annual_energy_col].fillna(0).sum()
    annual_electricity = selected_df[annual_elec_col].fillna(0).sum()
    annual_fuels = selected_df[annual_fuels_col].fillna(0).sum()
    annual_steam = selected_df[annual_steam_col].fillna(0).sum()

    selected_df[description_col] = clean_text(selected_df[description_col])

    temp_energy_df = pd.DataFrame({
        "Temperature Raw": selected_df[process_temp_web_col],
        "Annual Energy AU": selected_df[annual_energy_col],
    }).dropna(subset=["Temperature Raw", "Annual Energy AU"])

    temp_energy_df["Annual Energy Magnitude"] = temp_energy_df["Annual Energy AU"].abs()
    temp_energy_df = temp_energy_df[temp_energy_df["Annual Energy Magnitude"] > 0].copy()

    if not temp_energy_df.empty:
        temp_energy_df["Temperature Level"] = pd.cut(
            temp_energy_df["Temperature Raw"],
            bins=[-float("inf"), 20, 100, 200, 400, 600, float("inf")],
            labels=["<20 °C", "20-100 °C", "100-200 °C", "200-400 °C", "400-600 °C", ">=600 °C"],
            right=False
        )

        temp_breakdown_df = (
            temp_energy_df.groupby("Temperature Level", observed=False)["Annual Energy Magnitude"]
            .sum()
            .reset_index()
            .rename(columns={"Annual Energy Magnitude": "Value"})
        )

        temp_breakdown_df = temp_breakdown_df[temp_breakdown_df["Value"] > 0].copy()
        temp_breakdown_df["Temperature Level"] = temp_breakdown_df["Temperature Level"].astype(str)
    else:
        temp_breakdown_df = pd.DataFrame(columns=["Temperature Level", "Value"])

    detail_df = selected_df[
        [
            l3_col,
            description_col,
            naics_col,
            sec_elec_col,
            sec_fuels_col,
            sec_steam_col,
            efficiency_col,
            process_temp_col,
            process_temp_web_col,
            annual_energy_col,
            inlet_temp_col,
            outlet_temp_col,
            process_pressure_col,
            inlet_pressure_col,
            outlet_pressure_col,
            residence_time_col,
        ]
    ].rename(columns={
        l3_col: "Industry Application",
        description_col: "Description",
        naics_col: "NAICS Code",
        sec_elec_col: "SEC Electricity (GJ/t)",
        sec_fuels_col: "SEC Fuels (GJ/t)",
        sec_steam_col: "SEC Fuels or Electricity for Steam or Steam from CHP (GJ/t)",
        efficiency_col: "Efficiency (%)",
        process_temp_col: "Process temperature123 (°C)",
        process_temp_web_col: "Process Temperature (°C)",
        annual_energy_col: "Annual Energy from AU",
        inlet_temp_col: "Inlet temperature (°C)",
        outlet_temp_col: "Outlet temperature (°C)",
        process_pressure_col: "Process pressure (bar)",
        inlet_pressure_col: "Inlet pressure (bar)",
        outlet_pressure_col: "Outlet pressure (bar)",
        residence_time_col: "Residence time (sec)",
    })

    return {
        "Annual Production": annual_production,
        "Annual Energy": annual_energy,
        "Annual Electricity": annual_electricity,
        "Annual Fuels": annual_fuels,
        "Annual Steam": annual_steam,
        "Temperature Breakdown": temp_breakdown_df,
        "Rows": len(selected_df),
        "Details": detail_df,
        "Temperature Source Column": process_temp_web_col,
        "Annual Energy Source Column": "AU",
    }


def convert_fact_sheet_units(fact_sheet: dict, unit_system: str) -> dict:
    converted = fact_sheet.copy()

    if unit_system == "SI":
        return converted

    converted["Annual Energy"] = pj_to_tbtu(converted["Annual Energy"])
    converted["Annual Electricity"] = pj_to_tbtu(converted["Annual Electricity"])
    converted["Annual Fuels"] = pj_to_tbtu(converted["Annual Fuels"])
    converted["Annual Steam"] = pj_to_tbtu(converted["Annual Steam"])

    temp_breakdown = converted["Temperature Breakdown"].copy()
    if not temp_breakdown.empty:
        temp_breakdown["Value"] = temp_breakdown["Value"].apply(
            lambda x: pj_to_tbtu(x) if pd.notna(x) else x
        )
        temp_breakdown["Temperature Level"] = temp_breakdown["Temperature Level"].map(
            lambda x: TEMP_LABEL_MAP_SI_TO_IMPERIAL.get(str(x), str(x))
        )
        temp_breakdown = temp_breakdown[temp_breakdown["Value"] > 0].copy()
    converted["Temperature Breakdown"] = temp_breakdown

    details = converted["Details"].copy()

    sec_cols = [
        "SEC Electricity (GJ/t)",
        "SEC Fuels (GJ/t)",
        "SEC Fuels or Electricity for Steam or Steam from CHP (GJ/t)"
    ]
    for col in sec_cols:
        if col in details.columns:
            details[col] = details[col].apply(
                lambda x: gj_per_t_to_mmbtu_per_short_ton(x) if pd.notna(x) else x
            )

    temp_cols = [
        "Process temperature123 (°C)",
        "Process Temperature (°C)",
        "Inlet temperature (°C)",
        "Outlet temperature (°C)"
    ]
    for col in temp_cols:
        if col in details.columns:
            details[col] = details[col].apply(
                lambda x: c_to_f(x) if pd.notna(x) else x
            )

    pressure_cols = [
        "Process pressure (bar)",
        "Inlet pressure (bar)",
        "Outlet pressure (bar)"
    ]
    for col in pressure_cols:
        if col in details.columns:
            details[col] = details[col].apply(
                lambda x: bar_to_psi(x) if pd.notna(x) else x
            )

    if "Annual Energy from AU" in details.columns:
        details["Annual Energy from AU"] = details["Annual Energy from AU"].apply(
            lambda x: pj_to_tbtu(x) if pd.notna(x) else x
        )

    details = details.rename(columns={
        "SEC Electricity (GJ/t)": "SEC Electricity (MMBtu/short ton)",
        "SEC Fuels (GJ/t)": "SEC Fuels (MMBtu/short ton)",
        "SEC Fuels or Electricity for Steam or Steam from CHP (GJ/t)": "SEC Fuels or Electricity for Steam or Steam from CHP (MMBtu/short ton)",
        "Process temperature123 (°C)": "Process temperature123 (°F)",
        "Process Temperature (°C)": "Process Temperature (°F)",
        "Inlet temperature (°C)": "Inlet temperature (°F)",
        "Outlet temperature (°C)": "Outlet temperature (°F)",
        "Process pressure (bar)": "Process pressure (psi)",
        "Inlet pressure (bar)": "Inlet pressure (psi)",
        "Outlet pressure (bar)": "Outlet pressure (psi)",
        "Annual Energy from AU": "Annual Energy from AU (TBtu/yr)",
    })

    converted["Details"] = details
    return converted

# ----------------------------
# Chart builders
# ----------------------------
def build_top10_treemap(df: pd.DataFrame, title_root: str):
    fig = px.treemap(
        df,
        path=[px.Constant(title_root), "Unit operation (Level 2 classification)"],
        values="Display Percent",
        color="Display Percent",
        color_continuous_scale=["#D9EFEF", "#0B6E74"]
    )

    fig.update_traces(
        texttemplate="<b>%{label}</b><br>%{value:.1f}%",
        hovertemplate="<b>%{label}</b><br>Summed demand: %{value:.2f}%<extra></extra>",
        root_color="rgba(0,0,0,0)"
    )

    fig.update_layout(
        height=520,
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        margin=dict(t=40, l=20, r=20, b=20),
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=13),
        coloraxis_showscale=False
    )

    return fig


def build_annual_energy_donut(fact_sheet: dict, unit_system: str):
    annual_energy_unit = UNIT_CONFIG[unit_system]["annual_energy_unit"]

    donut_df = pd.DataFrame({
        "Energy Type": ["Electricity", "Fuels", "Steam"],
        "Value": [
            fact_sheet["Annual Electricity"],
            fact_sheet["Annual Fuels"],
            fact_sheet["Annual Steam"]
        ]
    })

    donut_df = donut_df[pd.to_numeric(donut_df["Value"], errors="coerce").fillna(0) > 0].copy()
    if donut_df.empty:
        return None

    fig = px.pie(
        donut_df,
        names="Energy Type",
        values="Value",
        hole=0.62,
        color="Energy Type",
        color_discrete_map=ENERGY_COLOR_MAP
    )

    total_energy = donut_df["Value"].sum()

    fig.update_traces(
        textposition="outside",
        texttemplate="%{label}<br>%{percent}",
        hovertemplate=(
            "<b>%{label}</b><br>"
            f"Value: %{{value:.3f}} {annual_energy_unit}<br>"
            "Share: %{percent}<extra></extra>"
        ),
        marker=dict(line=dict(color="#FFFFFF", width=2))
    )

    fig.update_layout(
        height=360,
        margin=dict(t=20, l=20, r=20, b=20),
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        showlegend=False,
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=13),
        annotations=[
            dict(
                text=f"<b>Total ({annual_energy_unit})</b><br>{total_energy:.2f}",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=16, color=TEXT_COLOR)
            )
        ]
    )

    return fig


def build_temperature_donut(fact_sheet: dict, unit_system: str):
    annual_energy_unit = UNIT_CONFIG[unit_system]["annual_energy_unit"]
    temp_color_map = UNIT_CONFIG[unit_system]["temp_color_map"]

    donut_df = fact_sheet["Temperature Breakdown"].copy()
    if donut_df.empty:
        return None

    donut_df["Value"] = pd.to_numeric(donut_df["Value"], errors="coerce")
    donut_df = donut_df.dropna(subset=["Temperature Level", "Value"])
    donut_df = donut_df[donut_df["Value"] > 0].copy()
    if donut_df.empty:
        return None

    donut_df["Temperature Level"] = donut_df["Temperature Level"].astype(str)

    fig = px.pie(
        donut_df,
        names="Temperature Level",
        values="Value",
        hole=0.62,
        color="Temperature Level",
        color_discrete_map=temp_color_map
    )

    total_energy = donut_df["Value"].sum()

    fig.update_traces(
        textposition="outside",
        texttemplate="%{label}<br>%{percent}",
        hovertemplate=(
            "<b>%{label}</b><br>"
            f"Energy magnitude: %{{value:.3f}} {annual_energy_unit}<br>"
            "Share: %{percent}<extra></extra>"
        ),
        marker=dict(line=dict(color="#FFFFFF", width=2))
    )

    fig.update_layout(
        height=360,
        margin=dict(t=20, l=20, r=20, b=20),
        paper_bgcolor=PAPER_BG,
        plot_bgcolor=PLOT_BG,
        showlegend=False,
        font=dict(family="Arial, sans-serif", color=TEXT_COLOR, size=13),
        annotations=[
            dict(
                text=f"<b>Total ({annual_energy_unit})</b><br>{total_energy:.2f}",
                x=0.5, y=0.5, showarrow=False,
                font=dict(size=16, color=TEXT_COLOR)
            )
        ]
    )

    return fig

# ----------------------------
# App UI
# ----------------------------
st.title("2022 U.S. Manufacturing Energy Consumption by Unit Operation and Energy Source")

unit_system = st.radio(
    "Select unit system",
    ["SI", "Imperial"],
    horizontal=True,
    key="unit_system"
)

try:
    df = load_excel_data(DATA_URL)

    all_l2_df = prepare_treemap_data(df, COL_PERCENT_ENERGY, top_n=999)
    overall_df = all_l2_df.nlargest(10, "Display Percent").reset_index(drop=True)
    electricity_df = prepare_treemap_data(df, COL_PERCENT_ELECTRICITY, top_n=10)
    fuels_df = prepare_treemap_data(df, COL_PERCENT_FUELS, top_n=10)
    steam_df = prepare_treemap_data(df, COL_PERCENT_STEAM, top_n=10)

    left_col, right_col = st.columns([1.6, 1.1], gap="large")

    with left_col:
        st.subheader("Total Energy Use Breakdown by Unit Operation (%)")
        st.plotly_chart(
            build_top10_treemap(overall_df, "Top 10 Categories by Total Energy Use"),
            use_container_width=True,
            theme=None,
            config={"displayModeBar": False, "scrollZoom": False}
        )
        st.caption("*Represented 2/3 of U.S. manufacturing sector in 2022.")

        st.subheader("Electricity Use Breakdown by Unit Operation (%)")
        st.plotly_chart(
            build_top10_treemap(electricity_df, "Top 10 Categories by Electricity Use"),
            use_container_width=True,
            theme=None,
            config={"displayModeBar": False, "scrollZoom": False}
        )
        st.caption("*Represented 2/3 of U.S. manufacturing sector in 2022.")
        st.subheader("Fuel Energy (Excluding Steam) Breakdown by Unit Operation (%)")
        st.plotly_chart(
            build_top10_treemap(fuels_df, "Top 10 Categories by Fuels (Excluding Steam) Use"),
            use_container_width=True,
            theme=None,
            config={"displayModeBar": False, "scrollZoom": False}
        )
        st.caption("*Represented 2/3 of U.S. manufacturing sector in 2022.")
        st.subheader("Steam Use Breakdown by Unit Operation (%)")
        st.plotly_chart(
            build_top10_treemap(steam_df, "Top 10 by Steam Use"),
            use_container_width=True,
            theme=None,
            config={"displayModeBar": False, "scrollZoom": False}
        )
        st.caption("*Represented 2/3 of U.S. manufacturing sector in 2022.")
    with right_col:
        selected_l2 = st.selectbox(
            "Select a unit operation to view its energy use breakdown",
            all_l2_df["Unit operation (Level 2 classification)"].tolist()
        )

        fact_sheet = build_fact_sheet(df, selected_l2)

        if fact_sheet is not None:
            display_fact_sheet = convert_fact_sheet_units(fact_sheet, unit_system)

            st.subheader("Annual Energy Use Breakdown")

            st.caption("Categorization by Energy Source")
            donut_fig = build_annual_energy_donut(display_fact_sheet, unit_system)
            if donut_fig is not None:
                st.plotly_chart(
                    donut_fig,
                    use_container_width=True,
                    theme=None,
                    config={"displayModeBar": False}
                )
            else:
                st.info("No positive annual energy values available for the selected category.")

            st.caption("Categorization by Process Temperature")
            temp_donut_fig = build_temperature_donut(display_fact_sheet, unit_system)
            if temp_donut_fig is not None:
                st.plotly_chart(
                    temp_donut_fig,
                    use_container_width=True,
                    theme=None,
                    config={"displayModeBar": False}
                )
            else:
                st.info(
                    f"No nonzero annual energy magnitudes from column "
                    f"{fact_sheet['Annual Energy Source Column']} with valid "
                    f"'{fact_sheet['Temperature Source Column']}' temperatures are available for the selected category."
                )

            detail_df = display_fact_sheet["Details"].drop(
                columns=[
                    "Process temperature123 (°C)",
                    "Process temperature123 (°F)",
                    "Annual Energy from AU",
                    "Annual Energy from AU (TBtu/yr)",
                    "Efficiency (%)",
                    "Inlet temperature (°C)",
                    "Inlet temperature (°F)",
                    "Outlet temperature (°C)",
                    "Outlet temperature (°F)",
                    "Inlet pressure (bar)",
                    "Inlet pressure (psi)",
                    "Outlet pressure (bar)",
                    "Outlet pressure (psi)",
                    "Residence time (sec)"
                ],
                errors="ignore"
            )

            st.dataframe(
                detail_df,
                use_container_width=True,
                hide_index=True
            )

except Exception as e:
    st.error(f"App error: {e}")
